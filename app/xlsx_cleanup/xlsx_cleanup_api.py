from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
import boto3
import tempfile
import os
import openpyxl
import requests

router = APIRouter(tags=["xlsx-cleanup"])


class CleanupRequest(BaseModel):
    s3_url: str
    upload_url: str | None = None
    replace: bool = False


def download_from_s3(s3_url: str, download_path: str):
    if s3_url.startswith("s3://"):
        # Parse s3://bucket/key
        path = s3_url[5:]
        bucket, key = path.split("/", 1)
        s3 = boto3.client("s3")
        s3.download_file(bucket, key, download_path)
    elif s3_url.startswith("http://") or s3_url.startswith("https://"):
        # Handle pre-signed S3 URLs via HTTPS download
        resp = requests.get(s3_url, stream=True)
        if resp.status_code != 200:
            raise ValueError(f"Failed to download from signed URL: {resp.status_code} {resp.text}")
        with open(download_path, "wb") as f:
            for chunk in resp.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
    else:
        raise ValueError("Invalid URL. Must start with s3://, http:// or https://")


def cleanup_xlsx(input_path: str, output_path: str):
    # Keep formulas/styles; don't evaluate to values
    wb = openpyxl.load_workbook(input_path, data_only=False)

    for sheet in wb.worksheets:
        # 1) Clear any AutoFilter (filtered-out rows often have hidden=True)
        try:
            if getattr(sheet, "auto_filter", None):
                sheet.auto_filter.ref = None
                sheet.auto_filter = None
            # Some files also set filterMode; turn it off
            if getattr(sheet.sheet_properties, "filterMode", None):
                sheet.sheet_properties.filterMode = False
        except Exception:
            pass

        # 2) Delete hidden rows
        hidden_rows = [idx for idx, row_dim in sheet.row_dimensions.items() if getattr(row_dim, "hidden", False)]
        for idx in sorted(hidden_rows, reverse=True):
            sheet.delete_rows(idx)

        # 3) Delete hidden columns
        hidden_cols = [letter for letter, col_dim in sheet.column_dimensions.items() if getattr(col_dim, "hidden", False)]
        # openpyxl uses numeric indexes for delete_cols, convert letter -> index
        from openpyxl.utils import column_index_from_string
        for letter in sorted(hidden_cols, key=lambda l: column_index_from_string(l), reverse=True):
            sheet.delete_cols(column_index_from_string(letter))

        # 4) After normalizing, delete rows/columns that are completely blank
        #    (all cells None after cleanup).
        blank_rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
            if all(cell.value is None for cell in row):
                blank_rows.append(row_idx)
        for idx in sorted(blank_rows, reverse=True):
            sheet.delete_rows(idx)

        blank_cols = []
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            col_cells = [sheet.cell(row=r, column=col_idx).value for r in range(1, sheet.max_row + 1)]
            if all(v is None for v in col_cells):
                blank_cols.append(col_idx)
        for idx in sorted(blank_cols, reverse=True):
            sheet.delete_cols(idx)

        # 5) IMPORTANT: Do not touch merged cells / named ranges / tables outside of necessary deletions

        # 5) Normalize text cells only (no changes to numbers/dates/formulas)
        for row in sheet.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str):
                    val = v.strip()
                    if val in ("", "-", "N/A", "Not Applicable", "None"):
                        cell.value = None
                    else:
                        cell.value = val

    wb.save(output_path)


@router.post("/cleanup_xlsx")
def cleanup_xlsx_endpoint(request: CleanupRequest):
    import logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger("xlsx-cleanup")
    try:
        logger.info(f"Raw request model: {request.dict()}")
        logger.info(f"Received cleanup request: s3_url={request.s3_url}, replace={request.replace}")
        # Use a persistent /tmp directory for debugging instead of auto-cleanup
        tmpdir = tempfile.mkdtemp(prefix="xlsx_cleanup_")
        download_path = os.path.join(tmpdir, "input.xlsx")
        output_path = os.path.join(tmpdir, "cleaned.xlsx")
        logger.info(f"Temporary directory created (will persist for debugging): {tmpdir}")

        try:
            logger.info("Starting download_from_s3...")
            download_from_s3(request.s3_url, download_path)
            logger.info(f"Downloaded file saved at {download_path}")
        except Exception as e:
            logger.exception("Download from S3 failed")
            raise

        try:
            logger.info("Starting cleanup_xlsx...")
            cleanup_xlsx(download_path, output_path)
            logger.info(f"Cleanup completed. Clean file saved at {output_path}")
        except Exception as e:
            logger.exception("Cleanup process failed")
            raise

        if not request.upload_url:
            logger.error("upload_url not provided in request. Cannot upload cleaned file.")
            raise HTTPException(status_code=400, detail="upload_url is required for upload")

        # Upload via signed URL only
        logger.info(f"Uploading cleaned file via presigned URL")
        with open(output_path, "rb") as f:
            put_resp = requests.put(request.upload_url, data=f)
        if put_resp.status_code not in (200, 201):
            logger.error(f"Upload via signed URL failed: {put_resp.status_code} {put_resp.text}")
            raise HTTPException(status_code=500, detail="Upload via signed URL failed")
        logger.info("Upload via signed URL complete.")
        return {"cleaned_file_url": request.upload_url}
    except Exception as e:
        logger.exception("Cleanup endpoint failed with exception (traceback included)")
        raise HTTPException(status_code=500, detail=str(e))