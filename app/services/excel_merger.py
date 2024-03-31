import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection

class ExcelMerger:
    def __init__(self, template_path, data_path, output_path):
        self.template_path = template_path
        self.data_path = data_path
        self.output_path = output_path

        # Validate existence of template and data files
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file does not exist: {self.template_path}")
        if not os.path.exists(self.data_path):
            raise FileNotFoundError(f"Data file does not exist: {self.data_path}")


    def _copy_cell_style(self, source_cell, target_cell):
        target_cell.font = Font(**vars(source_cell.font))
        target_cell.border = Border(**vars(source_cell.border))
        target_cell.fill = PatternFill(**vars(source_cell.fill))
        target_cell.number_format = source_cell.number_format
        target_cell.protection = Protection(**vars(source_cell.protection))
        target_cell.alignment = Alignment(**vars(source_cell.alignment))

    def _copy_sheet_content(self, source_sheet, target_sheet):
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = target_sheet.cell(row=cell.row, column=cell.column)
                if cell.data_type == 'f':
                    new_cell.data_type = 'f'
                    new_cell.value = cell.value
                else:
                    new_cell.value = cell.value
                if cell.has_style:
                    self._copy_cell_style(cell, new_cell)
                if cell.row == 1:
                    col_letter = get_column_letter(cell.column)
                    target_sheet.column_dimensions[col_letter].width = source_sheet.column_dimensions[col_letter].width

    def copy_formulas_and_data(self):
        print(f"Merging {self.template_path}, {self.data_path} to {self.output_path}")
        template_wb = load_workbook(self.template_path)
        data_wb = load_workbook(self.data_path)
        output_wb = Workbook()
        output_wb.remove(output_wb.active)

        for sheet_name in template_wb.sheetnames:
            template_sheet = template_wb[sheet_name]
            output_sheet = output_wb.create_sheet(sheet_name)
            self._copy_sheet_content(template_sheet, output_sheet)

        for sheet_name in data_wb.sheetnames:
            if sheet_name in output_wb.sheetnames:
                del output_wb[sheet_name]
            data_sheet = data_wb[sheet_name]
            output_sheet = output_wb.create_sheet(sheet_name)
            self._copy_sheet_content(data_sheet, output_sheet)

        output_wb.save(self.output_path)

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: python script.py <template_path> <data_path> <output_path>")
        sys.exit(1)
    
    merger = ExcelMerger(sys.argv[1], sys.argv[2], sys.argv[3])
    merger.copy_formulas_and_data()
